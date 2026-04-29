"""
Excel Workbook Builder — Clinic No-Show Dashboard
Sheets: Executive Summary, SQL Queries, Analysis Tables, High-Risk Segments,
        Recommendations, Raw Data
"""
import pickle, sqlite3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# ── Load ───────────────────────────────────────────────────────────────────
with open("query_results.pkl", "rb") as f:
    results = pickle.load(f)

conn = sqlite3.connect("/home/claude/clinic_v2/clinic.db")
df_raw = pd.read_sql_query("SELECT * FROM appointments", conn)
conn.close()

# ── Palette ────────────────────────────────────────────────────────────────
C = dict(
    navy="1E3A5F", blue="2563EB", teal="0D9488", red="DC2626",
    orange="EA580C", amber="D97706", green="16A34A",
    light="F0F4F8", white="FFFFFF", alt="EBF5FB",
    hdr_txt="FFFFFF", muted="64748B"
)

def ff(size=10, bold=False, color="1A1A1A", italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def fill(hex_c): return PatternFill("solid", fgColor=hex_c)

def bdr(style="thin", color="D1D5DB"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Helper: write a styled data table ─────────────────────────────────────
def write_table(ws, df, row, col, title=None, rate_cols=None, col_widths=None):
    rate_cols = rate_cols or []
    if title:
        c = ws.cell(row=row, column=col, value=title)
        c.font = ff(11, bold=True, color=C["navy"])
        c.alignment = lft()
        row += 1

    headers = list(df.columns)
    for ci, h in enumerate(headers):
        c = ws.cell(row=row, column=col+ci, value=h.replace("_"," ").title())
        c.font    = ff(9, bold=True, color=C["hdr_txt"])
        c.fill    = fill(C["navy"])
        c.alignment = ctr()
        c.border  = bdr()

    for ri, rec in enumerate(df.itertuples(index=False), 1):
        row_fill = fill(C["alt"]) if ri % 2 == 0 else fill(C["white"])
        for ci, val in enumerate(rec):
            c = ws.cell(row=row+ri, column=col+ci, value=val)
            c.font      = ff(9)
            c.fill      = row_fill
            c.alignment = ctr()
            c.border    = bdr()

    # colour-scale on rate column
    for rc in rate_cols:
        if rc in headers:
            ci = headers.index(rc)
            col_letter = get_column_letter(col+ci)
            rng = f"{col_letter}{row+1}:{col_letter}{row+len(df)}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B"))

    if col_widths:
        for offset, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(col+offset)].width = w

    return row + len(df) + 2


# ══════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ══════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────
# SHEET 1: Executive Summary
# ─────────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 2
for col in "BCDEFGHIJKLMNO":
    ws1.column_dimensions[col].width = 14

# Banner
ws1.merge_cells("B2:O2")
ws1.row_dimensions[2].height = 52
b = ws1["B2"]
b.value  = "Clinic Appointment No-Show Dashboard"
b.font   = ff(22, bold=True, color=C["white"])
b.fill   = fill(C["navy"])
b.alignment = ctr()

ws1.merge_cells("B3:O3")
ws1.row_dimensions[3].height = 22
sub = ws1["B3"]
sub.value = "5,000 Appointments · Jan 2023 – Dec 2024  |  Factors linked to patient no-shows"
sub.font  = ff(10, color=C["muted"], italic=True)
sub.fill  = fill("F8FAFC")
sub.alignment = ctr()

# KPI Cards row 5-6
ws1.row_dimensions[5].height = 15
ws1.row_dimensions[6].height = 52

kpis = [
    ("B6:C6",  "Total Appointments", "5,000",    C["navy"]),
    ("D6:E6",  "No-Show Rate",       "21.86%",   C["red"]),
    ("F6:G6",  "Total No-Shows",     "1,093",    C["orange"]),
    ("H6:I6",  "Cancellation Rate",  "9.30%",    C["amber"]),
    ("J6:K6",  "Avg Waiting Days",   "10.3 days",C["teal"]),
    ("L6:M6",  "SMS Coverage",       "62.5%",    C["blue"]),
    ("N6:O6",  "Showed Rate",        "68.84%",   C["green"]),
]
for cell_range, label, value, color in kpis:
    ws1.merge_cells(cell_range)
    c = ws1[cell_range.split(":")[0]]
    c.value = f"{label}\n{value}"
    c.font  = Font(name="Calibri", size=12, bold=True, color=C["white"])
    c.fill  = fill(color)
    c.alignment = ctr()
    c.border = Border(
        left=Side(style="medium", color=C["white"]),
        right=Side(style="medium", color=C["white"]),
        top=Side(style="medium", color=C["white"]),
        bottom=Side(style="medium", color=C["white"])
    )

# Section header
ws1.row_dimensions[8].height = 22
ws1["B8"].value = "Key Analysis Tables"
ws1["B8"].font  = ff(12, bold=True, color=C["navy"])

row = 9
# Age group
row = write_table(ws1, results["noshow_by_age_group"]["df"], row, 2,
                  title="No-Show Rate by Age Group",
                  rate_cols=["no_show_rate"],
                  col_widths=[14,20,12,14])

# Waiting days
row = write_table(ws1, results["noshow_by_waiting_days"]["df"], row, 2,
                  title="No-Show Rate by Waiting Period",
                  rate_cols=["no_show_rate"],
                  col_widths=[14,20,12,14])

# Right side: SMS + Status distribution
write_table(ws1, results["noshow_by_sms"]["df"], 9, 8,
            title="SMS Reminder Impact",
            rate_cols=["no_show_rate"],
            col_widths=[14,20,12,14])

write_table(ws1, results["status_distribution"]["df"], 17, 8,
            title="Show Status Distribution",
            col_widths=[14,12,14])

# Day of week table
write_table(ws1, results["noshow_by_day"]["df"], row, 2,
            title="No-Show Rate by Day of Week",
            rate_cols=["no_show_rate"],
            col_widths=[14,20,12,14])

# Prev no-shows
write_table(ws1, results["noshow_by_previous_ns"]["df"], row, 8,
            title="Effect of Previous No-Shows",
            rate_cols=["no_show_rate"],
            col_widths=[18,20,12,14])


# ─────────────────────────────────────────────────────────────────────────
# SHEET 2: SQL Queries
# ─────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("SQL Queries")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 2
ws2.column_dimensions["B"].width = 90

ws2.merge_cells("B2:B2")
ws2.row_dimensions[2].height = 40
h = ws2["B2"]
h.value = "SQL Queries — Clinic No-Show Analysis"
h.font  = ff(16, bold=True, color=C["white"])
h.fill  = fill(C["navy"])
h.alignment = ctr()

qrow = 4
for key, v in results.items():
    ws2.row_dimensions[qrow].height = 20
    t = ws2.cell(row=qrow, column=2, value=v["title"])
    t.font = ff(11, bold=True, color=C["navy"])
    t.alignment = lft()
    qrow += 1
    for line in v["sql"].split("\n"):
        ws2.row_dimensions[qrow].height = 15
        c = ws2.cell(row=qrow, column=2, value=line)
        c.font  = Font(name="Courier New", size=9, color="1E3A5F")
        c.fill  = fill("F1F5F9")
        c.alignment = lft()
        qrow += 1
    qrow += 2


# ─────────────────────────────────────────────────────────────────────────
# SHEET 3: Charts Data + Charts
# ─────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Charts")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 2
for col in "BCDEFGHIJKLMNO":
    ws3.column_dimensions[col].width = 16

ws3.merge_cells("B2:O2")
ws3.row_dimensions[2].height = 36
h3 = ws3["B2"]
h3.value = "Visual Analysis — No-Show Rates"
h3.font  = ff(16, bold=True, color=C["white"])
h3.fill  = fill(C["navy"])
h3.alignment = ctr()

# Write data tables that charts will reference
r = 4

# --- Age group data for chart ---
age_df = results["noshow_by_age_group"]["df"][["age_group","no_show_rate"]]
ws3.cell(r, 2, "Age Group").font = ff(9, bold=True, color=C["white"])
ws3.cell(r, 2).fill = fill(C["navy"]); ws3.cell(r, 2).alignment = ctr()
ws3.cell(r, 3, "No-Show %").font = ff(9, bold=True, color=C["white"])
ws3.cell(r, 3).fill = fill(C["navy"]); ws3.cell(r, 3).alignment = ctr()
for i, row_d in enumerate(age_df.itertuples(index=False), 1):
    ws3.cell(r+i, 2, row_d[0]).alignment = ctr()
    ws3.cell(r+i, 3, row_d[1]).alignment = ctr()
age_start = r; age_end = r + len(age_df)

# --- Appt type data ---
atype_df = results["noshow_by_appt_type"]["df"][["Appointment_Type","no_show_rate"]]
at_r = r + len(age_df) + 3
ws3.cell(at_r, 2, "Appointment Type").font = ff(9, bold=True, color=C["white"])
ws3.cell(at_r, 2).fill = fill(C["navy"]); ws3.cell(at_r, 2).alignment = ctr()
ws3.cell(at_r, 3, "No-Show %").font = ff(9, bold=True, color=C["white"])
ws3.cell(at_r, 3).fill = fill(C["navy"]); ws3.cell(at_r, 3).alignment = ctr()
for i, row_d in enumerate(atype_df.itertuples(index=False), 1):
    ws3.cell(at_r+i, 2, row_d[0]).alignment = ctr()
    ws3.cell(at_r+i, 3, row_d[1]).alignment = ctr()
at_start = at_r; at_end = at_r + len(atype_df)

# --- Day of week ---
day_df = results["noshow_by_day"]["df"][["Appointment_Day","no_show_rate"]]
day_r = at_end + 3
ws3.cell(day_r, 2, "Day").font = ff(9, bold=True, color=C["white"])
ws3.cell(day_r, 2).fill = fill(C["navy"]); ws3.cell(day_r, 2).alignment = ctr()
ws3.cell(day_r, 3, "No-Show %").font = ff(9, bold=True, color=C["white"])
ws3.cell(day_r, 3).fill = fill(C["navy"]); ws3.cell(day_r, 3).alignment = ctr()
for i, row_d in enumerate(day_df.itertuples(index=False), 1):
    ws3.cell(day_r+i, 2, row_d[0]).alignment = ctr()
    ws3.cell(day_r+i, 3, row_d[1]).alignment = ctr()
day_start = day_r; day_end = day_r + len(day_df)

# --- SMS ---
sms_df = results["noshow_by_sms"]["df"][["SMS_Reminder","no_show_rate"]]
sms_r = day_end + 3
ws3.cell(sms_r, 2, "SMS Reminder").font = ff(9, bold=True, color=C["white"])
ws3.cell(sms_r, 2).fill = fill(C["navy"]); ws3.cell(sms_r, 2).alignment = ctr()
ws3.cell(sms_r, 3, "No-Show %").font = ff(9, bold=True, color=C["white"])
ws3.cell(sms_r, 3).fill = fill(C["navy"]); ws3.cell(sms_r, 3).alignment = ctr()
for i, row_d in enumerate(sms_df.itertuples(index=False), 1):
    ws3.cell(sms_r+i, 2, row_d[0]).alignment = ctr()
    ws3.cell(sms_r+i, 3, row_d[1]).alignment = ctr()
sms_start = sms_r; sms_end = sms_r + len(sms_df)

# --- Monthly volume ---
vol_df = results["volume_by_month"]["df"][["month","total_appointments","no_shows","cancellations"]]
vol_r = sms_end + 3
for ci, col_name in enumerate(["Month","Total","No-Shows","Cancelled"], 2):
    ws3.cell(vol_r, ci, col_name).font = ff(9, bold=True, color=C["white"])
    ws3.cell(vol_r, ci).fill = fill(C["navy"]); ws3.cell(vol_r, ci).alignment = ctr()
for i, row_d in enumerate(vol_df.itertuples(index=False), 1):
    for ci, val in enumerate(row_d, 2):
        ws3.cell(vol_r+i, ci, val).alignment = ctr()
vol_start = vol_r; vol_end = vol_r + len(vol_df)

# --- Status distribution for pie ---
stat_df = results["status_distribution"]["df"]
stat_r = vol_end + 3
for ci, col_name in enumerate(["Status","Count","Pct"], 2):
    ws3.cell(stat_r, ci, col_name).font = ff(9, bold=True, color=C["white"])
    ws3.cell(stat_r, ci).fill = fill(C["navy"]); ws3.cell(stat_r, ci).alignment = ctr()
for i, row_d in enumerate(stat_df.itertuples(index=False), 1):
    for ci, val in enumerate(row_d, 2):
        ws3.cell(stat_r+i, ci, val).alignment = ctr()
stat_start = stat_r; stat_end = stat_r + len(stat_df)

# ─ Charts ──────────────────────────────────────────────────────────────
def bar_chart(title, cats_ref, vals_ref, chart_col="E", chart_row=4, width=14, height=10):
    ch = BarChart()
    ch.type = "col"; ch.grouping = "clustered"
    ch.title = title; ch.y_axis.title = "No-Show %"
    ch.legend = None; ch.style = 10
    ch.width = width; ch.height = height
    data = Reference(ws3, min_col=vals_ref[0], min_row=vals_ref[1], max_row=vals_ref[2])
    cats = Reference(ws3, min_col=cats_ref[0], min_row=cats_ref[1]+1, max_row=cats_ref[2])
    ch.add_data(data, titles_from_data=False)
    ch.set_categories(cats)
    return ch

# Chart 1: Age group bar
ch1 = bar_chart("No-Show Rate by Age Group",
                (2, age_start, age_end), (3, age_start, age_end))
ws3.add_chart(ch1, "E4")

# Chart 2: Appointment type bar
ch2 = bar_chart("No-Show Rate by Appointment Type",
                (2, at_start, at_end), (3, at_start, at_end))
ws3.add_chart(ch2, "P4")

# Chart 3: Day of week bar
ch3 = bar_chart("No-Show Rate by Day of Week",
                (2, day_start, day_end), (3, day_start, day_end))
ws3.add_chart(ch3, "E22")

# Chart 4: SMS bar
ch4 = bar_chart("No-Show Rate by SMS Reminder",
                (2, sms_start, sms_end), (3, sms_start, sms_end), width=10, height=10)
ws3.add_chart(ch4, "P22")

# Chart 5: Monthly line chart
ch5 = LineChart()
ch5.title = "Appointment Volume Over Time"
ch5.y_axis.title = "Count"
ch5.width = 28; ch5.height = 12; ch5.style = 10
data5 = Reference(ws3, min_col=3, max_col=5, min_row=vol_start, max_row=vol_end)
cats5 = Reference(ws3, min_col=2, min_row=vol_start+1, max_row=vol_end)
ch5.add_data(data5, titles_from_data=True)
ch5.set_categories(cats5)
ws3.add_chart(ch5, "E38")

# Chart 6: Pie chart — status
ch6 = PieChart()
ch6.title = "Show vs No-Show vs Cancelled"
ch6.width = 14; ch6.height = 12; ch6.style = 10
data6 = Reference(ws3, min_col=3, min_row=stat_start, max_row=stat_end)
cats6 = Reference(ws3, min_col=2, min_row=stat_start+1, max_row=stat_end)
ch6.add_data(data6, titles_from_data=False)
ch6.set_categories(cats6)
slice_colors = ["1E8449","C0392B","E67E22"]
for i, hex_c in enumerate(slice_colors):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = hex_c
    ch6.series[0].dPt.append(pt)
ws3.add_chart(ch6, "P38")


# ─────────────────────────────────────────────────────────────────────────
# SHEET 4: Analysis Tables (All query results full)
# ─────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Analysis Tables")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 2
for col in "BCDEFGHIJKL":
    ws4.column_dimensions[col].width = 18

ws4.merge_cells("B2:L2")
ws4.row_dimensions[2].height = 36
h4 = ws4["B2"]
h4.value = "Complete Analysis Tables"
h4.font  = ff(16, bold=True, color=C["white"])
h4.fill  = fill(C["navy"])
h4.alignment = ctr()

tab_row = 4
for key, v in results.items():
    if key in ("overall_noshow_rate",): continue
    df_t = v["df"]
    rate_col = [c for c in df_t.columns if "rate" in c.lower() or "pct" in c.lower() or "percentage" in c.lower()]
    tab_row = write_table(ws4, df_t, tab_row, 2,
                          title=v["title"],
                          rate_cols=rate_col,
                          col_widths=[max(14, len(str(c))+2) for c in df_t.columns])


# ─────────────────────────────────────────────────────────────────────────
# SHEET 5: High-Risk Segments
# ─────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("High-Risk Segments")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 2
for col in "BCDEFG":
    ws5.column_dimensions[col].width = 20

ws5.merge_cells("B2:G2")
ws5.row_dimensions[2].height = 36
h5 = ws5["B2"]
h5.value = "Top 10 Highest-Risk Patient Segments"
h5.font  = ff(16, bold=True, color=C["white"])
h5.fill  = fill(C["red"])
h5.alignment = ctr()

ws5["B4"].value = "These patient-clinic-reminder combinations carry the highest no-show risk."
ws5["B4"].font  = ff(10, color=C["muted"], italic=True)

write_table(ws5, results["high_risk_segments"]["df"], 6, 2,
            rate_cols=["no_show_rate"],
            col_widths=[12, 22, 16, 10, 16])

# Context box
ws5.merge_cells("B20:G24")
ctx = ws5["B20"]
ctx.value = (
    "Interpretation: All top-10 high-risk segments share one trait — no SMS reminder was sent. "
    "The 31-45 / Dermatology / No-Reminder segment leads at 42.86%. "
    "Young patients (0-17) without reminders appear in 4 of the top 10 slots across multiple clinic types, "
    "suggesting guardians/parents need targeted outreach as well."
)
ctx.font = ff(9, color="1E3A5F")
ctx.fill = fill("EFF6FF")
ctx.alignment = Alignment(wrap_text=True, vertical="top")
ctx.border = bdr("medium", "2563EB")


# ─────────────────────────────────────────────────────────────────────────
# SHEET 6: Recommendations
# ─────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Recommendations")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions["A"].width = 2
ws6.column_dimensions["B"].width = 6
ws6.column_dimensions["C"].width = 36
ws6.column_dimensions["D"].width = 60

ws6.merge_cells("B2:D2")
ws6.row_dimensions[2].height = 36
h6 = ws6["B2"]
h6.value = "5 Recommendations to Reduce No-Shows"
h6.font  = ff(15, bold=True, color=C["white"])
h6.fill  = fill(C["teal"])
h6.alignment = ctr()

recs = [
    ("#", "Action", "Evidence & Detail"),
    ("1", "Send SMS reminders to patients with longer wait times",
     "No-reminder patients have a 28.11% no-show rate vs 18.11% for reminded patients — a 10 pp gap. "
     "Patients waiting 15–30 days have the highest no-show rate (27.56%). Targeting long-wait patients "
     "with a reminder at booking AND 48h before the appointment can save an estimated 100+ appointments per month."),
    ("2", "Prioritise follow-up reminders for patients with previous no-shows",
     "Patients with 3+ prior no-shows have a 40.34% no-show rate — double the baseline. "
     "Implement a flag in the booking system so staff are alerted when a high-history patient books, "
     "triggering a personal phone call in addition to SMS."),
    ("3", "Avoid scheduling high-risk appointment types too far in advance",
     "Lab Tests (23.76%) and Consultations (23.73%) have the highest no-show rates by type. "
     "Where clinically safe, compress lead time to under 7 days for these categories. "
     "Same-day and 1-7 day bookings hold at only ~19-20% no-show."),
    ("4", "Add confirmation calls for appointments booked more than 14 days ahead",
     "No-show rate jumps from ~20% (1-14 days wait) to 27.56% at 15-30 days and 29.43% at 30+ days. "
     "A brief confirmation call at the 7-day mark reconfirms intent and allows rebooking, "
     "opening the slot for another patient if needed."),
    ("5", "Monitor no-show trends by day of week to improve staffing and scheduling",
     "Monday has the highest no-show rate (23.01%), 4 pp above Tuesday's low (19.02%). "
     "Deliberate overbooking on Mondays (10-15% buffer) and incentivised mid-week rescheduling "
     "can smooth out demand. Weekly tracking dashboards should flag day-of-week anomalies in real time."),
]

for ri, (num, title, detail) in enumerate(recs, 3):
    is_hdr = ri == 3
    ws6.row_dimensions[ri + 1].height = 75 if not is_hdr else 22
    cn = ws6.cell(row=ri+1, column=2, value=num)
    ct = ws6.cell(row=ri+1, column=3, value=title)
    cd = ws6.cell(row=ri+1, column=4, value=detail)
    for c in (cn, ct, cd):
        c.border  = bdr()
        c.alignment = Alignment(horizontal="center" if c is cn else "left",
                                vertical="center", wrap_text=True)
    if is_hdr:
        for c in (cn, ct, cd):
            c.font = ff(9, bold=True, color=C["white"])
            c.fill = fill(C["navy"])
    else:
        bg = C["alt"] if ri % 2 == 0 else C["white"]
        cn.font = ff(16, bold=True, color=C["navy"]); cn.fill = fill(bg)
        ct.font = ff(10, bold=True, color=C["navy"]); ct.fill = fill(bg)
        cd.font = ff(9, color="374151");              cd.fill = fill(bg)


# ─────────────────────────────────────────────────────────────────────────
# SHEET 7: Raw Data (sample 1000)
# ─────────────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("Raw Data (Sample)")
ws7.sheet_view.showGridLines = False

sample = df_raw.sample(1000, random_state=42).reset_index(drop=True)
cols = list(sample.columns)
col_widths_raw = [14,12,6,8,16,16,14,16,16,18,18,14,20,14,14,12]
for ci, (col_name, w) in enumerate(zip(cols, col_widths_raw), 1):
    ws7.column_dimensions[get_column_letter(ci)].width = w
    c = ws7.cell(row=1, column=ci, value=col_name.replace("_"," ").title())
    c.font = ff(9, bold=True, color=C["white"])
    c.fill = fill(C["navy"])
    c.alignment = ctr()
    c.border = bdr()

for ri, rec in enumerate(sample.itertuples(index=False), 2):
    bg = C["alt"] if ri % 2 == 0 else C["white"]
    for ci, val in enumerate(rec, 1):
        c = ws7.cell(row=ri, column=ci, value=val)
        c.font = ff(8); c.fill = fill(bg)
        c.alignment = ctr(); c.border = bdr()

# ─────────────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────────────
out = "Clinic_NoShow_Dashboard_v2.xlsx"
wb.save(out)
print(f"Saved: {out}")
